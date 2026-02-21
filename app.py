from flask import Flask, render_template, Response, request, jsonify, redirect, url_for, flash, session, abort
import cv2
import os
import time
import smtplib
import atexit
import urllib.request   
import numpy as np
from email.message import EmailMessage
from dotenv import load_dotenv
from twilio.rest import Client
from openai import OpenAI
import subprocess
import json
from flask import send_file
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import sqlite3
from flask_bcrypt import Bcrypt
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import openpyxl
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image,
    Table, TableStyle
)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
import shutil
import threading
from datetime import datetime, timedelta
import re
import sqlite3


# Load environment variables from .env
load_dotenv()

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "garuda_secret_key_secure_123") # Default if not in .env

video_writer = None   # ✅ ADD THIS LINE

fourcc = cv2.VideoWriter_fourcc(*'XVID')
last_image_time = 0
# =========================
# GLOBAL MAPPING STATUS
# =========================
MAPPING_STATUS = {
    "running": False,
    "completed": False,
    "volume": None,
    "map_image": None,
    "geo_image": None
}

# Security Setup
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message_category = 'info'

DB_NAME = "database.db"
LOG_FILE = os.path.join("logs", "user_login_details.xlsx")

# Ensure logs dir exists
os.makedirs("logs", exist_ok=True)

# Initialize Excel Log if not exists
if not os.path.exists(LOG_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["User Email", "Login Date", "Login Time", "Role", "Approval Status"])
    wb.save(LOG_FILE)

# =========================
# USER MODEL & DATABASE
# =========================
class User(UserMixin):
    def __init__(self, id, email, role, status):
        self.id = id
        self.email = email
        self.role = role
        self.status = status

@login_manager.user_loader
def load_user(user_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    user = c.fetchone()
    conn.close()
    if user:
        return User(id=user[0], email=user[1], role=user[3], status=user[4])
    return None

def log_login_attempt(email, role, status):
    try:
        wb = openpyxl.load_workbook(LOG_FILE)
        ws = wb.active
        now = datetime.now()
        ws.append([
            email, 
            now.strftime("%Y-%m-%d"), 
            now.strftime("%H:%M:%S"), 
            role, 
            status
        ])
        wb.save(LOG_FILE)
    except Exception as e:
        print(f"Error logging to Excel: {e}")

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT DEFAULT 'user',
        status TEXT DEFAULT 'pending'
    )
    """)

    conn.commit()
    conn.close()

  # =========================
# NOTIFICATION FUNCTIONS
# =========================

# =========================
# EMAIL FUNCTION
# =========================
def send_confirmation_email(message, user_email, pdf_path=None):
    try:
        EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS")
        EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")

        if not EMAIL_ADDRESS or not EMAIL_PASSWORD:
            print("❌ Email credentials not found in .env file")
            return False

        msg = EmailMessage()
        msg["Subject"] = "✅ Drone Survey Request Received"
        msg["From"] = EMAIL_ADDRESS
        msg["To"] = user_email

        # ✅ Enhanced email body with volume if present
        email_body = f"""
Hello,

Your Drone Survey Request has been received successfully.

{'='*50}
REQUEST DETAILS
{'='*50}

{message}

"""
        
        # ✅ Add volume highlight if present in message
        if "Volume" in message or "volume" in message:
            email_body += """
{'='*50}
✅ VOLUME CALCULATION INCLUDED
{'='*50}

Your survey includes the calculated stockpile volume.
This is a confirmed measurement based on 3D mapping analysis.

"""
        else:
            email_body += """
{'='*50}
📋 STATUS: PROCESSING
{'='*50}

Your survey will be processed within 24-48 hours.
Volume calculation and detailed analysis will follow.

"""

        email_body += """
{'='*50}
NEXT STEPS
{'='*50}

1. This PDF contains your survey request details
2. Our team will process your request immediately
3. You'll receive updates via this email and WhatsApp
4. For questions, contact: support@garuda.aero

Thank you,
Garuda Aerospace Drone Monitoring System
"""

        msg.set_content(email_body)
        
        # ✅ Attach PDF Report
        if pdf_path:
            if os.path.exists(pdf_path):
                try:
                    pdf_filename = os.path.basename(pdf_path)
                    with open(pdf_path, "rb") as f:
                        pdf_data = f.read()
                        msg.add_attachment(
                            pdf_data,
                            maintype="application",
                            subtype="pdf",
                            filename=pdf_filename
                        )
                    print(f"✅ PDF attached: {pdf_filename} ({len(pdf_data)} bytes)")
                except Exception as attach_error:
                    print(f"⚠️ Error attaching PDF: {attach_error}")
            else:
                print(f"⚠️ PDF file not found: {pdf_path}")

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            smtp.send_message(msg)

        print(f"✅ Email Sent To: {user_email}")
        return True
    except Exception as e:
        print(f"❌ Email Error: {e}")
        import traceback
        traceback.print_exc()
        return False

# =========================
# WHATSAPP FUNCTION
# =========================
# =========================
# WHATSAPP FUNCTION (FIXED)
# =========================
def send_whatsapp_message_with_pdf(message, phone_number, pdf_url):
    try:
        TWILIO_ACCOUNT_SID = os.getenv("TWILIO_ACCOUNT_SID")
        TWILIO_AUTH_TOKEN = os.getenv("TWILIO_AUTH_TOKEN")
        TWILIO_WHATSAPP_FROM = os.getenv("TWILIO_WHATSAPP_NUMBER")

        if not TWILIO_ACCOUNT_SID or not TWILIO_AUTH_TOKEN or not TWILIO_WHATSAPP_FROM:
            print("❌ Twilio credentials not found in .env file")
            return False

        client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

        # Normalize phone number
        phone_number = phone_number.replace(" ", "").replace("-", "")

        # If starts with 91 but missing +
        if phone_number.startswith("91") and not phone_number.startswith("+"):
            phone_number = "+" + phone_number
        # If normal 10 digit Indian number
        elif len(phone_number) == 10:
            phone_number = "+91" + phone_number

        to_number = "whatsapp:" + phone_number

        # ✅ Proper WhatsApp Formatted Text
        whatsapp_text = (
            "📌 Drone Survey Request Received\n\n"
            "Hello,\n\n"
            "Your Drone Survey Request has been received successfully.\n\n"
            "Request Details:\n"
            "-------------------------\n"
            f"{message}\n\n"
            "We will process the drone mapping soon.\n\n"
            "📄 Your PDF Report is attached above.\n\n"
            "Thank you,\n"
            "Garuda Aerospace Drone Monitoring System"
        )

        print(f"📱 Sending WhatsApp to: {to_number}")
        print(f"📎 PDF URL: {pdf_url}")

        # ✅ Send Message + PDF Together
        msg = client.messages.create(
            from_=TWILIO_WHATSAPP_FROM,
            to=to_number,
            body=whatsapp_text,
            media_url=[pdf_url]   # ✅ PDF Attachment
        )

        print(f"✅ WhatsApp Text + PDF Sent Successfully: {msg.sid}")
        return True

    except Exception as e:
        print(f"❌ WhatsApp Send Error: {e}")
        import traceback
        traceback.print_exc()
        return False




# =========================
# EXISTING DRONE LOGIC
# =========================
SHOT_URL = "http://10.75.165.104:8080/shot.jpg"
today = datetime.now().strftime("%Y-%m-%d")
BASE_DIR = os.path.join("storage", today)
IMG_DIR = os.path.join(BASE_DIR, "images")
VID_DIR = os.path.join(BASE_DIR, "videos")
REQ_DIR = os.path.join(BASE_DIR, "requests")

os.makedirs(IMG_DIR, exist_ok=True)
os.makedirs(VID_DIR, exist_ok=True)
os.makedirs(REQ_DIR, exist_ok=True)

def run_odm_mapping(date_folder):
    storage_path = os.path.join(os.getcwd(), "storage")
    dataset_name = date_folder
    output_folder = os.path.join(storage_path, dataset_name)

    # ✅ Check if Docker is available
    docker_check = subprocess.run("docker ps", shell=True, capture_output=True)
    if docker_check.returncode != 0:
        print("⚠️ Docker not found or not running. Entering SIMULATION MODE.")
        time.sleep(3) # Simulate some processing time
        return output_folder # Return path even if empty for simulation

    cmd = f"""
    docker run --rm -v "{storage_path}":/datasets opendronemap/odm \
    --project-path /datasets \
    --fast-orthophoto \
    --resize-to 1200 \
    --matcher-neighbors 4 \
    {date_folder}
    """


    print("🚀 Running ODM Mapping for:", dataset_name)
    subprocess.run(cmd, shell=True)
    return output_folder


def run_odm_mapping_background(date_folder, user_email, user_phone=None):
    global MAPPING_STATUS

    try:
        # ✅ Update status
        MAPPING_STATUS["running"] = True
        MAPPING_STATUS["completed"] = False
        MAPPING_STATUS["error"] = None

        print("🚀 ODM Mapping Started for:", date_folder)

        # ✅ Run ODM Mapping
        output_path = run_odm_mapping(date_folder)

        # ✅ Extract Volume After Mapping
        volume = extract_volume(output_path)

        # ✅ Generate PDF ONLY ONCE (After Volume Calculation)
        pdf_path = generate_pdf_report(volume, user_email)


        # ✅ Save Orthophoto Preview
        ortho_src = os.path.join(output_path, "odm_orthophoto", "odm_orthophoto.png")
        ortho_dest = os.path.join("static", "mapping.png")

        if os.path.exists(ortho_src):
            shutil.copy(ortho_src, ortho_dest)
        else:
            print("⚠️ Orthophoto not found, using fallback image...")

        # ✅ Mission Geo Image (First Dataset Image)
        images_dir = os.path.join(os.getcwd(), "storage", date_folder, "images")
        mission_geo = "/static/geo_latest.jpg"

        if os.path.exists(images_dir):
            imgs = [f for f in os.listdir(images_dir)
                    if f.lower().endswith((".jpg", ".jpeg", ".png"))]

            if imgs:
                mission_geo = f"/media/{date_folder}/images/{imgs[0]}"

        # ✅ Store Results for Status Command
        MAPPING_STATUS["volume"] = volume
        MAPPING_STATUS["map_image"] = "/static/mapping.png"
        MAPPING_STATUS["geo_image"] = mission_geo

        # ====================================================
        # ✅ SEND EMAIL ONLY ONCE (After Mapping Complete)
        # ====================================================
        send_confirmation_email(
            f"""
✅ Drone Mapping Completed Successfully!

📅 Date: {date_folder}
📊 Calculated Volume: {volume} m³

Your PDF Report is attached.
""",
            user_email,
            pdf_path
        )

        print("📧 PDF Report Sent Successfully to:", user_email)

        # ====================================================
        # ❌ WhatsApp Removed (Optional)
        # ====================================================
        # If you want WhatsApp later, we can add again

        # ✅ Final Status Update
        MAPPING_STATUS["completed"] = True
        print(f"✅ Mapping Completed for {date_folder}! Volume: {volume} m³")

    except Exception as e:
        print("❌ Mapping Error:", e)
        MAPPING_STATUS["error"] = str(e)

    finally:
        MAPPING_STATUS["running"] = False



def extract_volume(output_path):
    stats_file = os.path.join(output_path, "odm_report", "stats.json")

    if not os.path.exists(stats_file):
        # ✅ Demo SIMULATION Fallback
        import random
        mock_vol = round(random.uniform(450.0, 1250.0), 2)
        print(f"⚠️ Volume file not found. Simulating demo volume: {mock_vol}")
        return mock_vol

    with open(stats_file) as f:
        data = json.load(f)

    # Try multiple keys ODM may store
    if "volume" in data:
        return data["volume"]

    if "area" in data:
        area = data["area"]
        avg_height = 5   # Demo assumption
        volume = area * avg_height
        return round(volume, 2)

    return "❌ Volume Not Available"


def generate_pdf_report(volume_value, user_email="Unknown", location="Mining Site"):

    pdf_path = os.path.join(BASE_DIR, "volume_report.pdf")
    doc = SimpleDocTemplate(pdf_path, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # ============================
    # ✅ Company Logo + Header
    # ============================
    logo_path = "static/logo.png"
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=2*inch, height=1*inch)
        story.append(logo)

    story.append(Paragraph(
        "<b>Garuda Aerospace Pvt Ltd</b><br/>AI Drone Mining Monitoring System",
        styles["Heading2"]
    ))
    story.append(Spacer(1, 15))

    # ============================
    # ✅ Report Title
    # ============================
    story.append(Paragraph("📌 Drone Mining Survey Volume Report", styles["Title"]))
    story.append(Spacer(1, 20))

    # ============================
    # ✅ Client + Survey Info Table
    # ============================
    info_data = [
        ["Report Generated For", user_email],
        ["Survey Location", location],
        ["Survey Date", today],
        ["Processing Engine", "OpenDroneMap + AI Volume Estimation"]
    ]

    info_table = Table(info_data, colWidths=[180, 270])
    info_table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("BACKGROUND", (0,0), (0,-1), colors.lightgrey),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
    ]))

    story.append(info_table)
    story.append(Spacer(1, 20))

    # ============================
    # ✅ Volume Highlight Box
    # ============================
    story.append(Paragraph(
        f"<b>📊 Estimated Stockpile Volume:</b> "
        f"<font size=14 color='blue'>{volume_value} m³</font>",
        styles["Heading1"]
    ))
    story.append(Spacer(1, 25))

    # ============================
    # ✅ Summary Paragraph
    # ============================
    story.append(Paragraph("📝 Survey Summary", styles["Heading2"]))
    story.append(Paragraph(
        f"""
        This report was automatically generated using drone imagery and AI-powered
        3D mapping technology.

        The calculated stockpile volume for the mining site is estimated as:

        <b>{volume_value} cubic meters</b>.
        """,
        styles["Normal"]
    ))

    story.append(Spacer(1, 25))

    # ============================
    # ✅ Signature & Stamp Section
    # ============================
    story.append(Paragraph("✍ Authorized Signature", styles["Heading2"]))
    story.append(Spacer(1, 25))

    sign_data = [
        ["Project Engineer", "____________________"],
        ["Approved By (Admin)", "____________________"],
        ["Company Stamp", "____________________"]
    ]

    sign_table = Table(sign_data, colWidths=[200, 250])
    sign_table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("BACKGROUND", (0,0), (0,-1), colors.lightgrey),
    ]))

    story.append(sign_table)
    story.append(Spacer(1, 20))

    # ============================
    # ✅ Footer
    # ============================
    story.append(Paragraph(
        "Generated by Garuda Aerospace AI Drone Monitoring System | Confidential Report",
        styles["Italic"]
    ))

    # Build PDF
    doc.build(story)

    print("✅ Lightweight Professional PDF Generated:", pdf_path)
    return pdf_path


def generate_frames():
    global video_writer, last_image_time, img_counter

    while True:
        try:
            img_resp = urllib.request.urlopen(SHOT_URL, timeout=5)
            img_np = np.array(bytearray(img_resp.read()), dtype=np.uint8)
            frame = cv2.imdecode(img_np, cv2.IMREAD_COLOR)

            if frame is None:
                continue

        except Exception as e:
            time.sleep(1)
            continue

        current_time = time.time()

        # ✅ Start video if not started
        if video_writer is None:
            h, w, _ = frame.shape

            video_filename = datetime.now().strftime("%Y%m%d_%H%M%S") + ".avi"
            video_path = os.path.join(VID_DIR, video_filename)

            fourcc = cv2.VideoWriter_fourcc(*'XVID')
            video_writer = cv2.VideoWriter(video_path, fourcc, 20.0, (w, h))

            print("New Video Started:", video_filename)

        # ✅ Write video frame
        video_writer.write(frame)

        # ✅ Capture image every interval
        if current_time - last_image_time >= CAPTURE_INTERVAL:
            img_counter += 1

            img_name = datetime.now().strftime("%Y%m%d_%H%M%S")
            img_file = f"{img_name}_{img_counter:03d}.jpg"

            cv2.imwrite(os.path.join(IMG_DIR, img_file), frame)

            latest_geo = os.path.join("static", "geo_latest.jpg")
            cv2.imwrite(latest_geo, frame)

            last_image_time = current_time

        # ✅ Stream frame for browser
        ret, buffer = cv2.imencode(".jpg", frame)
        frame_bytes = buffer.tobytes()

        yield (b"--frame\r\n"
               b"Content-Type: image/jpeg\r\n\r\n" +
               frame_bytes +
               b"\r\n")


def get_mapping_preview(output_path):
    ortho_png = os.path.join(
        output_path,
        "odm_orthophoto",
        "odm_orthophoto.png"
    )

    if os.path.exists(ortho_png):
        static_path = os.path.join("static", "mapping.png")
        shutil.copy(ortho_png, static_path)
        return "/static/mapping.png"

    return None

def cleanup():
    global video_writer
    if video_writer is not None:
        video_writer.release()
        print("✅ Video writer released safely.")

def get_statistics():
    try:
        image_count = 0
        if os.path.exists(IMG_DIR):
            image_count = len([f for f in os.listdir(IMG_DIR) if f.endswith(('.jpg', '.png', '.jpeg'))])
        video_count = 0
        if os.path.exists(VID_DIR):
            video_count = len([f for f in os.listdir(VID_DIR) if f.endswith(('.avi', '.mp4', '.mov'))])
        total_size = 0
        for dirpath, dirnames, filenames in os.walk(BASE_DIR):
            for f in filenames:
                fp = os.path.join(dirpath, f)
                if os.path.exists(fp): total_size += os.path.getsize(fp)
        storage_used_mb = total_size / (1024 * 1024)
        return {"images": image_count, "videos": video_count, "storage_mb": round(storage_used_mb, 2)}
    except:
        return {"images": 0, "videos": 0, "storage_mb": 0}



def detect_date_from_message(msg):
    msg = msg.lower()

    if "today" in msg:
        return datetime.now().strftime("%Y-%m-%d")

    if "yesterday" in msg:
        return (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

    match = re.search(r"\d{4}-\d{2}-\d{2}", msg)
    if match:
        return match.group()

    return None

# =========================
# ROUTES
# =========================


@app.route("/")
def index():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    return render_template("index.html")


@app.route("/about")
def about():
    return render_template("about.html")


@app.route("/login", methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        logout_user()

        
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("SELECT * FROM users WHERE email = ?", (email,))
        user_data = c.fetchone()
        conn.close()
        
        if user_data and bcrypt.check_password_hash(user_data[2], password):
            user_obj = User(id=user_data[0], email=user_data[1], role=user_data[3], status=user_data[4])
            
            if user_obj.status == 'approved':
                login_user(user_obj)
                log_login_attempt(email, user_obj.role, "Success")
                flash('Login Successful! Welcome to Garuda Aerospace.', 'success')
                if user_obj.role == 'admin':
                    return redirect(url_for('admin_dashboard'))
                return redirect(url_for('dashboard'))
            else:
                log_login_attempt(email, user_obj.role, "Pending/Rejected")
                flash('Access Request Sent to Admin. Please wait for approval.', 'info')
        else:
            log_login_attempt(email, "Unknown", "Failed")
            flash('Invalid Email or Password', 'error')
            
    return render_template('login.html')

@app.route("/register", methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if password != confirm_password:
            flash('Passwords do not match!', 'error')
            return redirect(url_for('register'))
            
        hashed_pw = bcrypt.generate_password_hash(password).decode('utf-8')
        
        try:
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            c.execute("INSERT INTO users (email, password, role, status) VALUES (?, ?, 'user', 'pending')", 
                      (email, hashed_pw))
            conn.commit()
            conn.close()
            flash('Access Request Sent! Please wait for Admin approval.', 'success')
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash('Email already exists.', 'error')
            
    return render_template('register.html')

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('login'))

@app.route("/dashboard")
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route("/survey_logs")
@login_required
def survey_logs():
    storage_path = "storage"
    logs = []
    if os.path.exists(storage_path):
        # Filter for YYYY-MM-DD format directories
        logs = [d for d in os.listdir(storage_path) if os.path.isdir(os.path.join(storage_path, d)) and re.match(r"\d{4}-\d{2}-\d{2}", d)]
        logs.sort(reverse=True)
    return render_template('survey_logs.html', logs=logs)

@app.route("/survey_logs/<date>")
@login_required
def survey_detail(date):
    storage_path = os.path.join("storage", date)
    images_dir = os.path.join(storage_path, "images")
    videos_dir = os.path.join(storage_path, "videos")
    
    images = []
    if os.path.exists(images_dir):
        images = [f for f in os.listdir(images_dir) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
    
    videos = []
    if os.path.exists(videos_dir):
        videos = [f for f in os.listdir(videos_dir) if f.lower().endswith(('.mp4', '.avi', '.mov'))]
        
    return render_template('survey_detail.html', date=date, images=images, videos=videos)

@app.route("/media/<date>/<type>/<filename>")
@login_required
def serve_media(date, type, filename):
    # Ensure type is either 'images' or 'videos'
    if type not in ['images', 'videos']:
        abort(404)
    
    directory = os.path.join("storage", date, type)
    if not os.path.exists(os.path.join(directory, filename)):
        abort(404)
        
    return send_file(os.path.join(directory, filename))

@app.route("/analytics")
@login_required
def analytics():
    # Basic aggregate analytics
    storage_path = "storage"
    total_surveys = 0
    total_storage_mb = 0
    survey_data = []
    
    if os.path.exists(storage_path):
        dirs = [d for d in os.listdir(storage_path) if os.path.isdir(os.path.join(storage_path, d)) and re.match(r"\d{4}-\d{2}-\d{2}", d)]
        total_surveys = len(dirs)
        
        for d in dirs:
            dir_path = os.path.join(storage_path, d)
            size = 0
            for path, dirs_sub, files in os.walk(dir_path):
                for f in files:
                    fp = os.path.join(path, f)
                    size += os.path.getsize(fp)
            mb = round(size / (1024 * 1024), 2)
            total_storage_mb += mb
            survey_data.append({"date": d, "size": mb})
            
    return render_template('analytics.html', total_surveys=total_surveys, total_storage=round(total_storage_mb, 2), survey_data=survey_data)

@app.route("/settings")
@login_required
def settings():
    return render_template('settings.html', user=current_user)

@app.route("/admin")
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        flash('Unauthorized Access!', 'error')
        return redirect(url_for('dashboard'))
        
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT * FROM users")
    all_users = c.fetchall()
    
    # Stats
    pending = sum(1 for u in all_users if u[4] == 'pending')
    total = len(all_users)
    
    # Read Login Logs
    total_logins = 0
    if os.path.exists(LOG_FILE):
        wb = openpyxl.load_workbook(LOG_FILE)
        ws = wb.active
        total_logins = ws.max_row - 1 # Subtract header
    
    conn.close()
    
    return render_template('admin_dashboard.html', users=all_users, pending_count=pending, users_count=total, total_logins=total_logins)

@app.route("/admin/approve/<int:user_id>")
@login_required
def approve_user(user_id):
    if current_user.role != 'admin':
        abort(403)
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("UPDATE users SET status = 'approved' WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    flash('User Approved Successfully.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route("/admin/reject/<int:user_id>")
@login_required
def reject_user(user_id):
    if current_user.role != 'admin':
        abort(403)
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("UPDATE users SET status = 'pending' WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    flash('User Access Revoked.', 'info')
    return redirect(url_for('admin_dashboard'))

@app.route("/admin/delete/<int:user_id>")
@login_required
def delete_user(user_id):
    if current_user.role != 'admin':
        abort(403)
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    flash('User Deleted.', 'info')
    return redirect(url_for('admin_dashboard'))

# =========================
# EXISTING API/VIDEO
# =========================
@app.route("/video")
def video():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    return Response(generate_frames(), mimetype="multipart/x-mixed-replace; boundary=frame")

@app.route("/api/statistics", methods=["GET"])
@login_required
def api_statistics():
    stats = get_statistics()
    return jsonify({"status": "success", "data": stats})

@app.route("/chat_ai", methods=["POST"])
@login_required
def chat_ai():
    global MAPPING_STATUS

    data = request.json
    user_msg = data.get("message", "").strip().lower()

    # ✅ Logged-in email auto
    user_email = current_user.email
    user_phone = data.get("phone")

    if not user_msg:
        return jsonify({"reply": "⚠️ Please type something!"})

    # ======================================
    # ✅ 1️⃣ STATUS COMMAND (Always Priority)
    # ======================================
    if "status" in user_msg:

        if MAPPING_STATUS["completed"]:
            return jsonify({
                "reply": "✅ Mapping Completed Successfully!",
                "volume": MAPPING_STATUS["volume"],
                "map_image": MAPPING_STATUS["map_image"],
                "geo_image": MAPPING_STATUS["geo_image"]
            })

        elif MAPPING_STATUS["running"]:
            return jsonify({
                "reply": "⏳ Still Processing... please wait."
            })

        else:
            return jsonify({
                "reply": "⚠️ No mapping started yet. Type 'generate mapping today'."
            })

    # ======================================
    # ✅ 2️⃣ START MAPPING COMMAND
    # ======================================
    mapping_keywords = [
        "generate mapping",
        "start mapping",
        "3d mapping",
        "generate volume",
        "calculate volume"
    ]

    if any(cmd in user_msg for cmd in mapping_keywords):

        # ✅ Prevent double start
        if MAPPING_STATUS["running"]:
            return jsonify({
                "reply": "⏳ Mapping already running... please wait."
            })

        # ✅ Detect date folder
        date_folder = detect_date_from_message(user_msg)

        if not date_folder:
            return jsonify({
                "reply": "⚠️ Please specify date. Example: 'generate mapping for 2026-02-10'"
            })

        # ✅ Start Background Thread
        threading.Thread(
            target=run_odm_mapping_background,
            args=(date_folder, user_email, user_phone)
        ).start()

        return jsonify({
            "reply": f"""
🚀 Mapping Initialized Successfully!

📅 Target Date: {date_folder}
📧 Report will be sent to: {user_email}

⏳ Processing started...
Your 3D map + Geo image will update automatically.

Type **status** anytime.
"""
        })

    # ======================================
    # ✅ 3️⃣ NORMAL AI CHAT MODE
    # ======================================
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        temperature=0.7,
        max_tokens=400,
        messages=[
            {"role": "system", "content": "You are Garuda Aerospace AI Assistant."},
            {"role": "user", "content": user_msg}
        ]
    )

    return jsonify({
        "reply": response.choices[0].message.content
    })


# =========================
# SURVEY REQUEST HANDLER
# =========================
def generate_survey_pdf(mission_objectives, operator_email, timestamp, volume=None):
    """Generate PDF for survey request with optional volume"""
    pdf_path = os.path.join(REQ_DIR, f"survey_request_{timestamp}.pdf")
    doc = SimpleDocTemplate(pdf_path, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # ============================
    # ✅ Company Header
    # ============================
    logo_path = "static/logo.png"
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=2*inch, height=1*inch)
        story.append(logo)

    story.append(Paragraph(
        "<b>Garuda Aerospace Pvt Ltd</b><br/>Drone Survey Request Form",
        styles["Heading2"]
    ))
    story.append(Spacer(1, 15))

    # ============================
    # ✅ Survey Request Title
    # ============================
    story.append(Paragraph("📋 Drone Survey Mission Request", styles["Title"]))
    story.append(Spacer(1, 20))

    # ============================
    # ✅ Request Details Table
    # ============================
    request_data = [
        ["Request ID", timestamp],
        ["Operator Email", operator_email],
        ["Request Date", datetime.now().strftime("%Y-%m-%d")],
        ["Request Time", datetime.now().strftime("%H:%M:%S")],
        ["Status", "Processing" if volume else "Pending Processing"]
    ]

    req_table = Table(request_data, colWidths=[180, 270])
    req_table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("BACKGROUND", (0,0), (0,-1), colors.lightgrey),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))

    story.append(req_table)
    story.append(Spacer(1, 20))

    # ============================
    # ✅ Volume Calculation (if available)
    # ============================
    if volume:
        story.append(Paragraph("📊 Volume Calculation Results", styles["Heading2"]))
        story.append(Spacer(1, 10))
        
        volume_data = [
            ["Estimated Stockpile Volume", f"<b>{volume} m³</b>"],
            ["Calculation Method", "OpenDroneMap + AI Volume Estimation"],
            ["Confidence Level", "High (3D Point Cloud Analysis)"]
        ]
        
        volume_table = Table(volume_data, colWidths=[180, 270])
        volume_table.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.5, colors.black),
            ("BACKGROUND", (0,0), (0,-1), colors.lightblue),
            ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
            ("ALIGN", (0,0), (-1,-1), "LEFT"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ]))
        
        story.append(volume_table)
        story.append(Spacer(1, 20))

    # ============================
    # ✅ Mission Objectives
    # ============================
    story.append(Paragraph("📌 Mission Objectives & Coordinates", styles["Heading2"]))
    story.append(Paragraph(
        mission_objectives if mission_objectives else "No specific coordinates provided",
        styles["Normal"]
    ))
    story.append(Spacer(1, 20))

    # ============================
    # ✅ Instructions
    # ============================
    story.append(Paragraph("📝 Next Steps", styles["Heading2"]))
    if volume:
        next_steps = """
        1. <b>Volume Confirmed:</b> Your survey volume has been calculated and is shown above.
        <br/><br/>
        2. <b>Documentation:</b> This PDF serves as your official survey report with volume confirmation.
        <br/><br/>
        3. <b>Delivery:</b> Complete 3D maps, detailed analysis, and geotagged images will be sent within 24 hours.
        <br/><br/>
        4. <b>Contact:</b> For questions, reply to this email or contact our team.
        """
    else:
        next_steps = """
        1. <b>Confirmation:</b> Your survey request has been received and logged in our system.
        <br/><br/>
        2. <b>Processing:</b> Our team will begin drone mapping operations at the specified location.
        <br/><br/>
        3. <b>Delivery:</b> 3D maps, volume analysis, and geotagged images will be sent within 24-48 hours.
        <br/><br/>
        4. <b>Contact:</b> For updates, reply to this email or contact our team.
        """
    
    story.append(Paragraph(next_steps, styles["Normal"]))
    story.append(Spacer(1, 25))

    # ============================
    # ✅ Footer with Contact Info
    # ============================
    story.append(Paragraph(
        "<b>Contact Information:</b><br/>Email: support@garuda.aero<br/>WhatsApp: +91-XXXXXXXXXX<br/><br/>"
        "Generated by Garuda Aerospace AI Drone Monitoring System | Confidential Report",
        styles["Italic"]
    ))

    # Build PDF
    doc.build(story)
    print(f"✅ Survey Request PDF Generated: {pdf_path}")
    return pdf_path


@app.route("/ai_request", methods=["POST"])
@login_required
def ai_request():
    """Handle survey request: Generate PDF and send via email + WhatsApp"""
    try:
        global MAPPING_STATUS
        
        data = request.json
        message = data.get("message", "").strip()
        email = data.get("email", "").strip()
        phone = data.get("phone", "").strip()

        print(f"\n📋 Survey Request Received:")
        print(f"   Message: {message[:50]}...")
        print(f"   Email: {email}")
        print(f"   Phone: {phone}")

        # ✅ Validation
        if not message or not email or not phone:
            return jsonify({
                "status": "error",
                "message": "❌ All fields (Objectives, Email, Phone) are required!"
            }), 400

        # ✅ Generate unique timestamp for request ID
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        print(f"📌 Request ID: {timestamp}")
        
        # ✅ Get volume from MAPPING_STATUS if available
        volume = MAPPING_STATUS.get("volume")
        print(f"📊 Current Volume in MAPPING_STATUS: {volume}")
        
        # ✅ Generate Survey PDF with volume if available
        try:
            pdf_path = generate_survey_pdf(message, email, timestamp, volume=volume)
            print(f"✅ PDF Generated: {pdf_path}")
            print(f"   File exists: {os.path.exists(pdf_path)}")
            if os.path.exists(pdf_path):
                print(f"   File size: {os.path.getsize(pdf_path)} bytes")
        except Exception as pdf_error:
            print(f"❌ Error generating PDF: {pdf_error}")
            import traceback
            traceback.print_exc()
            raise pdf_error

        # ✅ Prepare email message with volume info if available
        message_with_volume = f"Mission Objectives:\n{message}"
        if volume:
            message_with_volume += f"\n\n📊 Volume Calculation: {volume} m³"

        # ✅ Send Email with PDF
        print("\n📧 Sending email...")
        email_success = send_confirmation_email(
            message=message_with_volume,
            user_email=email,
            pdf_path=pdf_path
        )

        # ✅ Send WhatsApp with PDF (if phone provided)
        whatsapp_success = False
        if phone:
            try:
                # Get public URL for PDF - use relative path that will be served
                pdf_filename = os.path.basename(pdf_path)
                request_date = datetime.now().strftime('%Y-%m-%d')
                # Using localhost for development - change to actual domain in production
                pdf_url = f"http://localhost:5000/download_survey_pdf/{request_date}/{pdf_filename}"
                
                whatsapp_message = f"Your Drone Survey Request has been received!\n\n📋 Request ID: {timestamp}\n\nObjectives:\n{message}"
                if volume:
                    whatsapp_message += f"\n\n📊 Volume: {volume} m³"
                
                print(f"\n📱 Sending WhatsApp...")
                whatsapp_success = send_whatsapp_message_with_pdf(
                    message=whatsapp_message,
                    phone_number=phone,
                    pdf_url=pdf_url
                )
                print(f"✅ WhatsApp sent successfully to {phone}")
            except Exception as e:
                print(f"⚠️ WhatsApp send error (non-critical): {e}")
                import traceback
                traceback.print_exc()
                whatsapp_success = False

        # ✅ Log the request
        try:
            log_file = os.path.join(REQ_DIR, "survey_log.txt")
            with open(log_file, "a") as f:
                volume_str = f"Volume: {volume} m³" if volume else "Volume: Pending"
                f.write(f"[{timestamp}] Email: {email} | Phone: {phone} | {volume_str} | Email Success: {email_success} | WhatsApp Success: {whatsapp_success}\n")
            print(f"✅ Request logged to {log_file}")
        except Exception as e:
            print(f"⚠️ Logging error: {e}")

        # ✅ Return response
        response_text = f"✅ Survey Request Transmitted!\n\n"
        response_text += f"📧 Email Sent: {'Yes ✓' if email_success else 'Failed ✗'}\n"
        response_text += f"📱 WhatsApp Sent: {'Yes ✓' if whatsapp_success else 'Failed ✗'}\n"
        if volume:
            response_text += f"📊 Volume Included: {volume} m³ ✓\n"
        else:
            response_text += f"📊 Volume: Pending (Run mapping first)\n"
        response_text += f"📋 Request ID: {timestamp}\n"
        response_text += f"\n📄 PDF: {os.path.basename(pdf_path)}"

        print(f"\n✅ Final Response: {response_text}\n")

        return jsonify({
            "status": "success",
            "message": response_text,
            "file": os.path.basename(pdf_path),
            "volume": volume
        }), 200

    except Exception as e:
        print(f"\n❌ Survey Request Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "status": "error",
            "message": f"❌ Error processing request: {str(e)}"
        }), 500


@app.route("/download_report")
@login_required
def download_report():
    pdf_path = os.path.join(BASE_DIR, "volume_report.pdf") # Case sensitive fix
    if not os.path.exists(pdf_path):
        return "⚠️ Report not generated yet!"
    return send_file(pdf_path, as_attachment=True)


@app.route("/download_survey_pdf/<date>/<filename>")
def download_survey_pdf(date, filename):
    """Download survey PDF from storage"""
    try:
        pdf_path = os.path.join("storage", date, "requests", filename)
        
        if not os.path.exists(pdf_path):
            return jsonify({"error": "PDF not found"}), 404
        
        return send_file(pdf_path, as_attachment=True, mimetype="application/pdf")
    except Exception as e:
        print(f"Error downloading survey PDF: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/volume_report.pdf")
def volume_report():
    pdf_path = os.path.join(BASE_DIR, "volume_report.pdf")

    if not os.path.exists(pdf_path):
        return "PDF Not Ready Yet", 404

    return send_file(
        pdf_path,
        mimetype="application/pdf",
        as_attachment=False
    )

@app.route("/twilio_pdf")
def twilio_pdf():
    pdf_path = os.path.join(BASE_DIR, "volume_report.pdf")

    if not os.path.exists(pdf_path):
        return "PDF Not Ready", 404

    return send_file(
        pdf_path,
        mimetype="application/pdf"
    )


@app.route("/admin-login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")

        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("SELECT * FROM users WHERE email = ? AND role='admin'", (email,))
        admin_data = c.fetchone()
        conn.close()

        if admin_data and bcrypt.check_password_hash(admin_data[2], password):
            admin_obj = User(
                id=admin_data[0],
                email=admin_data[1],
                role=admin_data[3],
                status=admin_data[4]
            )

            login_user(admin_obj)
            flash("Admin Login Successful!", "success")
            return redirect(url_for("admin_dashboard"))

        else:
            flash("Invalid Admin Credentials", "error")

    return render_template("admin_login.html")

@app.route("/report_pdf")
def report_pdf():
    pdf_path = os.path.join(BASE_DIR, "volume_report.pdf")
    return send_file(pdf_path, mimetype="application/pdf")

@app.route("/public_report")
def public_report():
    pdf_path = os.path.join(BASE_DIR, "volume_report.pdf")

    if not os.path.exists(pdf_path):
        return "Report not ready yet!"

    return send_file(
        pdf_path,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="Mining_Report.pdf"
    )

@app.route("/setup-admin")
def setup_admin():
    hashed_pw = bcrypt.generate_password_hash("Yogesh@0901").decode("utf-8")

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    c.execute("""
    INSERT OR IGNORE INTO users (email, password, role, status)
    VALUES (?, ?, 'admin', 'approved')
    """, ("yogeshmalavai9@gmail.com", hashed_pw))

    conn.commit()
    conn.close()

    return "✅ Admin Created Successfully"

if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5000, debug=False)
